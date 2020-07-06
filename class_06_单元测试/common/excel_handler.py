#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @CreateTime    : 2020-07-04 11:34
# @Author  : guoDD
# @Email   : Email
# @File    : excel_handler
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

"""
1.打开表单
2.读取表单
"""


class ExcelHandler():
    """操作excel"""

    def __init__(self, file):
        """初始化函数"""
        self.file = file

    def open_sheet(self, name) -> Worksheet:
        """打开表单"""
        """在函数或者方法后面加 -> 类型:   表示此函数返回值是一个这样的类型"""
        """函数注解"""
        wb = load_workbook(self.file)
        sheet = wb[name]
        return sheet

    def get_header(self, sheet_name):
        """获取表单的表头  标题"""
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    # def read(self, sheet_name):
    #     """读取所有额数据"""
    #     sheet = self.open_sheet(sheet_name)
    #     rows = list(sheet.rows)[1:]
    #     data = []
    #     for row in rows:
    #         row_data = []
    #         for cell in row:
    #             row_data.append(cell.value)
    #         data.append(row_data)
    #     return data

    def read(self, sheet_name):
        """读取所有额数据"""
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)
        # 获取标题

        data = []
        for row in rows[1:]:
            row_data = []
            # 获取标题
            for cell in row:
                row_data.append(cell.value)
                # 列表转成字段，要和header去zip
                data_dict = dict(zip(self.get_header(sheet_name), row_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write(file, sheet_name, row, column, data):
        """写入excel数据"""
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(file)
        # 关闭
        wb.close()



if __name__ == '__main__':
    excel = ExcelHandler(r'e:\test.xlsx')
    sheet = excel.open_sheet("Sheet1")
    print(sheet)
    header = excel.get_header("Sheet1")
    print(header)
    data = excel.read("Sheet1")
    print(data)

    excel1 = ExcelHandler(r'e:\test.xlsx')
    excel1.write(r'e:\test.xlsx', "Sheet1", 4, 1, "guodongdong111")
