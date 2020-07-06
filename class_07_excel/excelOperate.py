#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @CreateTime    : 2020-07-03 17:22
# @Author  : guoDD
# @Email   : Email
# @File    : excelOperate

from openpyxl.reader.excel import load_workbook
import os


class ExcelMethod():

    def __init__(self, filename, sheetName):

        self.filename = filename
        self.wb = load_workbook(filename)
        # 通过工作的表名获取一个工作表对象
        self.sheet = self.wb[sheetName]
        # 获取工作表中的最大行号
        self.maxRowNum = self.sheet.max_row
        # 获取工作表中的最大列号
        self.max_column = self.sheet.max_column
        # 获取总的行
        self.row = self.sheet.max_row

    def read_excel(self):

        dataList = []
        try:
            for row in self.sheet.rows:
                tmpList = []
                for cell in row:
                    tmpList.append(cell.value)
                dataList.append(tmpList)
        except:
         print("%s加载失败" % self.filename)

        else:
         return dataList[1:]


    def write_excel(self, row, Column, text):
        try:
            self.sheet.cell(row, Column, text)

        except:
            print("%s 写入失败" % self.filename)

    def save_excel(self):
        try:
            self.wb.save(self.filename)
        except:
            print("%s 保存失败" % self.filename)

    def close_excel(self):
        try:
            self.wb.close()
        except:
            print("%s 关闭失败" % self.filename)

if __name__=="__main__":
    # dri_url = os.path.join(os.getcwd() + r"\test.xlsx")
    dri_url = r'e:\test.xlsx'
    excel = ExcelMethod(dri_url,"Sheet1")
    dataJson = excel.read_excel()
    excel.write_excel(5,3,"henghenghahei")
    excel.save_excel()
    print(dataJson)
    excel.close_excel()
