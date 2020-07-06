#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @CreateTime    : 2020-07-01 18:15
# @Author  : guoDD
# @Email   : Email
# @File    : demo_01_excel

"""
"""
import openpyxl

"""读取excel文件夹"""
# windows下的路径有反斜杠
file_name = r'e:\test.xlsx'
wb = openpyxl.load_workbook(file_name)
# print(wb)
# 不直接去获取_sheets属性，称为私有属性
# print(wb._sheets)    #self._sheets

# active是表示被激活的，被选中的sheet
active_sheet = wb.active

# sheetnames 和 _sheets 有什么区别？
# sheetnames列表当中存储的是字符串， _sheets里面存储的是对象

# 获取所有表单的正确用法
wb.worksheets
# print(wb.worksheets)

"""获取某一个表单"""
# 1.通过索引获取
# sheet = wb.worksheets[0]
# 2.通过表单名称获取sheet1
# sheet = wb.get_sheet_by_name("Sheet1")# 弃用
# 正规用法
sheet= wb['Sheet1']
# print(sheet)

# pycharm  不支持 sheet.属性 的提示

# 读取单个单元格
# 行和列是从1开始的，不是python当中从0开始
cell = sheet.cell(1,1)
# print(cell)
# 获取cell的值
# print(cell.value)

# 获取某一行
# print(sheet[1])
# 获取某一列
# print(sheet['A'])

# 获取多行
# print(sheet[1:3])
#
# for column in sheet[1]:
#     print(column.value)


# 获取所有的数据
total_data = list(sheet.rows)
print(total_data)

for row in total_data:
    for cell in row:
        print(cell.value)



#  写成类

# 写入，单元格

# 一定记得保存
wb.save(file_name)
# 关闭
wb.close()

"""
总结：
获取表单的方法
1.  wb.active,被选中被激活
2.通过索引，  wb.worksheets[索引]
3.通过sheet名字， wb[表单名称]
"""